"""商品修正依頼の修正前後比較Excelを生成する。"""

from pathlib import Path
import os
import re
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from product_requests import (
    ProductCorrectionLine,
    ProductCorrectionRequest,
    validate_correction_lines,
)


DEFAULT_EXPORT_DIRECTORY = Path(
    os.environ.get("REVISION_EXPORT_DIRECTORY", r"C:\tsv_auto\exports\修正前後比較")
)
HEADER_FILL = PatternFill("solid", fgColor="5B9BD5")
AFTER_FILL = PatternFill("solid", fgColor="E2F0D9")
WHITE_BOLD_FONT = Font(color="FFFFFF", bold=True)
THIN_GRAY_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)


def _fallback_source_values(line: ProductCorrectionLine) -> dict[str, str]:
    return {
        "自治体ID": line.product.municipality_id,
        "自治体名": line.product.municipality_name,
        "お礼の品ID": line.product.product_id,
        "オリジナルお礼の品ID": line.product.original_product_id,
        "（必須）お礼の品名": line.product.product_name,
        "事業者ID": line.product.business_id,
        "サイト表示事業者名": line.product.business_name,
    }


def _build_product_snapshots(
    lines: list[ProductCorrectionLine],
) -> tuple[list[str], list[list[str]], list[list[str]], set[tuple[int, int]]]:
    """元データと、指定列だけを差し替えた修正後データを組み立てる。"""

    headers: list[str] = []
    products: dict[tuple[str, str], dict[str, str]] = {}
    changes: dict[tuple[str, str], dict[str, str]] = {}
    for line in lines:
        key = (line.product.municipality_id, line.product.product_id)
        source_values = line.product.source_values() or _fallback_source_values(line)
        products.setdefault(key, source_values)
        changes.setdefault(key, {})[line.field_name] = line.after_value
        for header in source_values:
            if header not in headers:
                headers.append(header)
        if line.field_name not in headers:
            headers.append(line.field_name)

    original_rows = []
    revised_rows = []
    changed_cells: set[tuple[int, int]] = set()
    for row_index, (key, source_values) in enumerate(products.items(), start=2):
        original = [source_values.get(header, "") for header in headers]
        revised_values = dict(source_values)
        revised_values.update(changes[key])
        revised = [revised_values.get(header, "") for header in headers]
        original_rows.append(original)
        revised_rows.append(revised)
        for header in changes[key]:
            changed_cells.add((row_index, headers.index(header) + 1))
    return headers, original_rows, revised_rows, changed_cells


def _write_product_data_sheet(
    worksheet,
    headers: list[str],
    rows: list[list[str]],
    changed_cells: set[tuple[int, int]],
) -> None:
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_GRAY_BORDER
    worksheet.row_dimensions[1].height = 36
    for row in rows:
        worksheet.append(row)
    for row_number in range(2, worksheet.max_row + 1):
        for column_number in range(1, len(headers) + 1):
            cell = worksheet.cell(row=row_number, column=column_number)
            cell.border = THIN_GRAY_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_number, column_number in changed_cells:
        if row_number <= worksheet.max_row:
            worksheet.cell(row=row_number, column=column_number).fill = AFTER_FILL
    for column_number, header in enumerate(headers, start=1):
        content_lengths = [len(str(header))]
        content_lengths.extend(
            len(str(row[column_number - 1])) for row in rows if len(row) >= column_number
        )
        worksheet.column_dimensions[get_column_letter(column_number)].width = min(
            max(max(content_lengths) + 2, 12), 32
        )
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(headers))}{max(worksheet.max_row, 1)}"
    )
    worksheet.sheet_view.showGridLines = False


def _safe_file_name(value: str) -> str:
    return re.sub(r'[\\/:*?"<>|]+', "_", value)


def generate_revision_comparison_workbook(
    request: ProductCorrectionRequest,
    lines: Iterable[ProductCorrectionLine],
    output_directory: Path = DEFAULT_EXPORT_DIRECTORY,
) -> Path:
    """選択商品の変更後データだけを含むExcelを出力する。"""

    validated_lines = validate_correction_lines(request, lines)
    output_directory.mkdir(parents=True, exist_ok=True)
    output_path = output_directory / f"変更後データ_{_safe_file_name(request.request_id)}.xlsx"

    product_headers, original_rows, revised_rows, changed_cells = _build_product_snapshots(
        validated_lines
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "商品データ"
    _write_product_data_sheet(
        worksheet,
        product_headers,
        revised_rows,
        changed_cells,
    )

    workbook.save(output_path)
    return output_path
