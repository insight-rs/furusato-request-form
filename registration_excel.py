"""新規商品登録用Excelテンプレートの生成・取込を行う。"""

from __future__ import annotations

from dataclasses import dataclass
import base64
from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from config_master import normalize
from form_definitions import RequestFormField
from choice_reference import load_choice_categories, load_local_product_standards


TEMPLATE_ID = "FURUSATO-CHOICE-REGISTRATION"
TEMPLATE_VERSION = "1.0"
PRODUCT_SHAPES = ("単品", "定期便", "SKU展開")
LEGACY_TEMPLATE_PATH = Path(__file__).resolve().parent / "config" / "legacy_registration_template.b64"

NOULESS_REFERENCE_FIELDS = (
    ("商品代（税込）", "必須", ""),
    ("在庫数", "必須", ""),
    ("税率", "必須", "8%|10%|非課税"),
    ("JANコード", "任意", ""),
    ("参考サイトURL", "任意", ""),
    ("SNS URL", "任意", ""),
    ("集荷先名称", "任意", ""),
    ("集荷先郵便番号", "任意", ""),
    ("集荷先住所", "任意", ""),
    ("集荷先電話番号", "任意", ""),
    ("集荷先担当者", "任意", ""),
    ("集荷先担当者メール", "任意", ""),
    ("在庫方式", "必須", "売り切り|月間|無制限"),
    ("食品区分", "必須", "食品|食品以外"),
    ("原材料", "食品の場合必須", ""),
    ("食品一括表示ファイル名", "食品の場合任意", ""),
    ("掲載ストーリー：こだわり", "任意", ""),
    ("掲載ストーリー：人", "任意", ""),
    ("掲載ストーリー：場所", "任意", ""),
    ("掲載ストーリー：歴史", "任意", ""),
    ("掲載ストーリー：想い", "任意", ""),
    ("掲載ストーリー：ふるさと納税による変化", "任意", ""),
)


@dataclass(frozen=True)
class RegistrationExcelImport:
    product_shape: str
    choice_values: dict[str, str]
    extra_values: dict[str, str]
    subscription_rows: list[dict[str, str]]
    sku_rows: list[dict[str, str]]
    warnings: tuple[str, ...] = ()


def _section(field: RequestFormField) -> str:
    name = field.source_column
    if name.startswith("アレルギー：") or "アレルギー" in name:
        return "食品・アレルギー"
    if any(word in name for word in ("配送", "発送", "配達")):
        return "配送情報"
    if any(word in name for word in ("画像", "動画")):
        return "画像・動画"
    if any(word in name for word in ("表示", "受付", "限定")):
        return "公開・受付設定"
    if any(word in name for word in ("寄付", "寄附", "ポイント", "還元率")):
        return "金額・ポイント"
    return "商品情報"


def _style_sheet(sheet, widths: dict[str, float]) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _style_header(cells) -> None:
    fill = PatternFill("solid", fgColor="234A62")
    for cell in cells:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _add_list_validation(sheet, cell_ref: str, options: list[str]) -> None:
    clean = [str(option).replace('"', '""') for option in options if str(option)]
    formula = '"' + ",".join(clean) + '"'
    if clean and len(formula) <= 255:
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.error = "一覧から選択してください。"
        validation.errorTitle = "入力内容を確認してください"
        sheet.add_data_validation(validation)
        validation.add(cell_ref)


def _attach_original_application_form(workbook, choice_values, extra_values):
    """共通申請書の書式を保った先頭シートを追加し、認識可能な値を反映する。"""
    if not LEGACY_TEMPLATE_PATH.exists():
        return workbook
    legacy = load_workbook(BytesIO(base64.b64decode(LEGACY_TEMPLATE_PATH.read_text(encoding="ascii"))))
    form = legacy.active
    form.title = "新規返礼品申込フォーム"
    # 提供された記入例を空テンプレート化する。書式・結合・数式は維持する。
    input_cells = (
        "H6", "H8", "J10", "J11", "J13", "J15", "V15", "D18", "J21", "P21", "V21",
        "D22", "D24", "D28", "J28", "J31", "J32", "V32", "J34", "V34", "J36", "J38",
        "J39", "V39", "J41", "V41", "J43", "T43", "J46", "N46", "R46", "V46", "Y46",
        "J48", "N49", "V49", "J51", "N52", "V52", "N54", "T54", "J57", "D60", "H75",
    )
    for coordinate in input_cells:
        form[coordinate] = None
    for row in range(68, 96):
        form[f"E{row}"] = None

    def value(*columns):
        return next((normalize(choice_values.get(column, "")) for column in columns if normalize(choice_values.get(column, ""))), "")

    mapped = {
        "H6": value("site business"),
        "J10": value("管理コード"),
        "J13": value("（必須）お礼の品名"),
        "J15": value("（条件付き必須）寄附額", "寄附額"),
        "V15": normalize(extra_values.get("商品代（税込）", "")),
        "D18": value("容量"),
        "D22": value("キャッチコピー"),
        "D24": value("説明"),
        "J28": value("消費期限"),
        "J31": value("原材料名"),
        "J38": value("地場産品類型"),
        "J41": value("温度帯"),
        "J57": normalize(extra_values.get("在庫数", "")),
    }
    for coordinate, stored in mapped.items():
        if stored:
            form[coordinate] = stored

    # アレルギーは〇表記へ戻す。
    allergy_row_by_name = {
        normalize(form[f"C{row}"].value): row for row in range(68, 96)
        if normalize(form[f"C{row}"].value)
    }
    for column, stored in choice_values.items():
        if "アレルギー" not in column or normalize(stored) not in {"1", "あり", "○", "〇"}:
            continue
        name = column.split("：")[-1].replace("フラグ", "").strip("（）() ")
        if name in allergy_row_by_name:
            form[f"E{allergy_row_by_name[name]}"] = "〇"

    # 独自の全項目・定期便・SKUシートを後ろに付け、テンプレート外項目も欠落させない。
    for source in workbook.worksheets:
        target = legacy.create_sheet(source.title)
        target.sheet_format = copy(source.sheet_format)
        target.sheet_properties = copy(source.sheet_properties)
        target.freeze_panes = source.freeze_panes
        target.sheet_view.showGridLines = source.sheet_view.showGridLines
        for key, dimension in source.column_dimensions.items():
            target.column_dimensions[key] = copy(dimension)
        for key, dimension in source.row_dimensions.items():
            target.row_dimensions[key] = copy(dimension)
        for row in source.iter_rows():
            for cell in row:
                copied = target[cell.coordinate]
                copied.value = cell.value
                if cell.has_style:
                    copied._style = copy(cell._style)
                if cell.number_format:
                    copied.number_format = cell.number_format
                copied.alignment = copy(cell.alignment)
        for merged in source.merged_cells.ranges:
            target.merge_cells(str(merged))
        for validation in source.data_validations.dataValidation:
            target.add_data_validation(copy(validation))
        target.auto_filter.ref = source.auto_filter.ref
    return legacy


def build_registration_template(
    fields: Iterable[RequestFormField],
    product_shape: str,
    *,
    choice_values: dict[str, str] | None = None,
    extra_values: dict[str, str] | None = None,
    subscription_rows: list[dict[str, str]] | None = None,
    sku_rows: list[dict[str, str]] | None = None,
) -> bytes:
    """チョイス全列と商品形態別明細を持つ公式テンプレートを作る。"""

    if product_shape not in PRODUCT_SHAPES:
        raise ValueError(f"未対応の商品形態です: {product_shape}")
    fields = list(fields)
    choice_values = choice_values or {}
    extra_values = extra_values or {}
    workbook = Workbook()
    guide = workbook.active
    guide.title = "はじめに"
    guide["A1"] = "ふるさと納税 新規商品登録テンプレート"
    guide["A1"].font = Font(size=18, bold=True, color="234A62")
    guide["A2"], guide["B2"] = "テンプレートID", TEMPLATE_ID
    guide["A3"], guide["B3"] = "バージョン", TEMPLATE_VERSION
    guide["A4"], guide["B4"] = "商品形態", product_shape
    guide["A6"] = "入力方法"
    guide["A7"] = "黄色の入力欄へ記入してください。選択肢がある項目は一覧から選択できます。"
    guide["A8"] = "チョイス全項目シートは、ふるさとチョイスの商品マスタ列を順番どおり網羅しています。"
    guide["A9"] = "定期便・SKU展開の場合は、専用明細シートも入力してください。"
    _style_sheet(guide, {"A": 24, "B": 70})

    choice = workbook.create_sheet("チョイス全項目")
    headers = ["区分", "項目ID", "商品マスタ列名", "表示名", "必須区分", "入力形式", "入力値", "選択肢・説明"]
    choice.append(headers)
    _style_header(choice[1])
    yellow = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9E1E8")
    for row_number, field in enumerate(fields, start=2):
        options = [label for label, _ in field.options()]
        stored = normalize(choice_values.get(field.source_column, ""))
        label_by_code = {code: label for label, code in field.options()}
        choice.append([
            _section(field), field.field_id, field.source_column, field.label,
            field.requirement, field.input_kind, label_by_code.get(stored, stored),
            " / ".join(options) or field.source_instruction,
        ])
        choice.cell(row_number, 7).fill = yellow
        _add_list_validation(choice, f"G{row_number}", options)
        for cell in choice[row_number]:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    choice.column_dimensions["B"].hidden = True
    choice.column_dimensions["C"].hidden = True
    choice.column_dimensions["F"].hidden = True
    choice.auto_filter.ref = f"A1:H{choice.max_row}"
    _style_sheet(choice, {"A": 20, "B": 16, "C": 32, "D": 34, "E": 14, "F": 12, "G": 42, "H": 60})

    extras = workbook.create_sheet("共通追加情報")
    extras.append(["項目", "必須区分", "入力値", "選択肢・説明"])
    _style_header(extras[1])
    for row_number, (label, requirement, options_text) in enumerate(NOULESS_REFERENCE_FIELDS, start=2):
        extras.append([label, requirement, extra_values.get(label, ""), options_text.replace("|", " / ")])
        extras.cell(row_number, 3).fill = yellow
        _add_list_validation(extras, f"C{row_number}", options_text.split("|") if options_text else [])
    _style_sheet(extras, {"A": 36, "B": 18, "C": 48, "D": 55})

    category = workbook.create_sheet("カテゴリー選択")
    category.append(["カテゴリー番号", "大項目", "中項目", "小項目", "カテゴリーID（自動取込用）"])
    _style_header(category[1])
    categories = load_choice_categories()
    selected_category_ids = [
        value.strip() for value in normalize(choice_values.get("（必須）カテゴリー", "")).split("|")
        if value.strip()
    ]
    for index in range(3):
        selected = next(
            (row for row in categories if index < len(selected_category_ids) and row.category_id == selected_category_ids[index]),
            None,
        )
        category.append([
            index + 1,
            selected.major if selected else "",
            selected.middle if selected else "",
            selected.minor if selected else "",
            selected.category_id if selected else "",
        ])
        for cell in category[category.max_row][1:4]:
            cell.fill = yellow
    category["A6"] = "参照用カテゴリー一覧（大項目・中項目・小項目の順に選んでください）"
    category["A6"].font = Font(bold=True, color="234A62")
    for row in categories:
        category.append(["", row.major, row.middle, row.minor, row.category_id])
    _style_sheet(category, {"A": 18, "B": 28, "C": 32, "D": 38, "E": 24})

    local = workbook.create_sheet("地場産品類型")
    local.append(["選択コード", "地場産品に該当する理由"])
    _style_header(local[1])
    local_value = normalize(choice_values.get("地場産品類型", ""))
    local_code, _, local_reason = local_value.partition("|")
    local.append([local_code, local_reason])
    local["A2"].fill = yellow
    local["B2"].fill = yellow
    local["A4"] = "コード"
    local["B4"] = "類型の説明"
    _style_header(local[4])
    standards = load_local_product_standards()
    for code, description in standards:
        local.append([code, description])
    _add_list_validation(local, "A2", [code for code, _ in standards])
    _style_sheet(local, {"A": 18, "B": 85})

    if product_shape == "定期便":
        detail = workbook.create_sheet("定期便明細")
        detail_headers = ["お届け回", "お届け時期", "お届け内容", "内容量", "数量", "温度帯", "補足"]
        detail.append(detail_headers)
        _style_header(detail[1])
        for index, values in enumerate(subscription_rows or [{} for _ in range(12)], start=1):
            detail.append([values.get("お届け回", str(index))] + [values.get(header, "") for header in detail_headers[1:]])
            for cell in detail[detail.max_row]:
                cell.fill = yellow
        _add_list_validation(detail, "F2:F100", ["常温", "冷蔵", "冷凍"])
        _style_sheet(detail, {"A": 12, "B": 20, "C": 42, "D": 22, "E": 12, "F": 14, "G": 45})
    elif product_shape == "SKU展開":
        detail = workbook.create_sheet("SKU明細")
        detail_headers = [
            "登録区分", "商品区分", "既存商品", "品番取得方法", "SKU品番", "商品名", "バリエーション名",
            "品種", "容量", "色", "数量", "配送月", "その他の分け方",
            "商品代変更", "商品代（税込）", "寄附額", "在庫数", "温度帯", "補足",
        ]
        detail.append(detail_headers)
        _style_header(detail[1])
        for values in sku_rows or [{} for _ in range(30)]:
            detail.append([values.get(header, "") for header in detail_headers])
            for cell in detail[detail.max_row]:
                cell.fill = yellow
        _add_list_validation(detail, "B2:B200", ["新規", "既存"])
        _add_list_validation(detail, "D2:D200", ["品番を入力する", "品番取得を依頼する"])
        _add_list_validation(detail, "N2:N200", ["商品代を登録する", "商品代を変更する", "商品代を変更しない"])
        _add_list_validation(detail, "R2:R200", ["常温", "冷蔵", "冷凍"])
        _style_sheet(detail, {
            "A": 20, "B": 12, "C": 38, "D": 22, "E": 18, "F": 34, "G": 26,
            "H": 18, "I": 16, "J": 14, "K": 12, "L": 16, "M": 24,
            "N": 22, "O": 18, "P": 16, "Q": 14, "R": 14, "S": 42,
        })

        sku_master = workbook.create_sheet("SKUチョイスマスタ")
        field_list = list(fields)
        master_headers = ["SKU No.", "登録区分", "元商品区分"] + [field.source_column for field in field_list]
        sku_master.append(master_headers)
        _style_header(sku_master[1])
        for index, row in enumerate(sku_rows or [], start=1):
            inherited = row.get("チョイスマスタ値", {})
            if not isinstance(inherited, dict):
                inherited = {}
            master_values = dict(choice_values)
            if row.get("商品区分") == "既存":
                master_values.update({key: normalize(value) for key, value in inherited.items()})
            sku_code = normalize(row.get("SKU品番", ""))
            sku_name = normalize(row.get("商品名", ""))
            for column in ("管理コード", "連携コード"):
                if sku_code:
                    master_values[column] = sku_code
            if sku_name:
                master_values["（必須）お礼の品名"] = (
                    f"{sku_name} {sku_code}" if sku_code and not sku_name.endswith(sku_code) else sku_name
                )
            if row.get("寄附額"):
                master_values["（条件付き必須）必要寄付金額"] = normalize(row.get("寄附額"))
            if row.get("容量"):
                master_values["容量"] = normalize(row.get("容量"))
            temperature = normalize(row.get("温度帯", ""))
            for label, column in {
                "常温": "常温配送フラグ", "冷蔵": "冷蔵配送フラグ", "冷凍": "冷凍配送フラグ"
            }.items():
                if temperature:
                    master_values[column] = "1" if temperature == label else "0"
            sku_master.append([
                index, "新規登録（SKU）", normalize(row.get("商品区分", "")),
                *[master_values.get(field.source_column, "") for field in field_list],
            ])
        sku_master.freeze_panes = "D2"
        sku_master.auto_filter.ref = f"A1:{sku_master.cell(1, len(master_headers)).column_letter}{max(1, sku_master.max_row)}"
        _style_sheet(sku_master, {"A": 12, "B": 20, "C": 14})

    workbook = _attach_original_application_form(workbook, choice_values, extra_values)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def read_registration_template(content: bytes, fields: Iterable[RequestFormField]) -> RegistrationExcelImport:
    """公式テンプレートまたは旧申請書を読み、保存コードへ正規化する。"""

    workbook = load_workbook(BytesIO(content), data_only=False)
    if "はじめに" not in workbook.sheetnames:
        return _read_legacy_workbook(workbook, fields)
    guide = workbook["はじめに"]
    if normalize(guide["B2"].value) != TEMPLATE_ID:
        raise ValueError("対応していないExcelテンプレートです。")
    product_shape = normalize(guide["B4"].value) or "単品"
    field_by_column = {field.source_column: field for field in fields}
    choice_values: dict[str, str] = {}
    warnings = []
    if "チョイス全項目" in workbook.sheetnames:
        for row in workbook["チョイス全項目"].iter_rows(min_row=2, values_only=True):
            source_column = normalize(row[2] if len(row) > 2 else "")
            value = normalize(row[6] if len(row) > 6 else "")
            if not source_column or not value:
                continue
            field = field_by_column.get(source_column)
            if field is None:
                warnings.append(f"未登録のチョイス項目：{source_column}")
                continue
            choice_values[source_column] = dict(field.options()).get(value, value)
    extra_values: dict[str, str] = {}
    if "共通追加情報" in workbook.sheetnames:
        for row in workbook["共通追加情報"].iter_rows(min_row=2, values_only=True):
            label = normalize(row[0] if row else "")
            value = normalize(row[2] if len(row) > 2 else "")
            if label and value:
                extra_values[label] = value
    if "カテゴリー選択" in workbook.sheetnames:
        category_sheet = workbook["カテゴリー選択"]
        categories = load_choice_categories()
        selected_ids = []
        for row in category_sheet.iter_rows(min_row=2, max_row=4, values_only=True):
            major = normalize(row[1] if len(row) > 1 else "")
            middle = normalize(row[2] if len(row) > 2 else "")
            minor = normalize(row[3] if len(row) > 3 else "")
            category_id = normalize(row[4] if len(row) > 4 else "")
            if not category_id and major:
                matches = [
                    item for item in categories
                    if item.major == major
                    and (not middle or item.middle == middle)
                    and (not minor or item.minor == minor)
                ]
                if matches:
                    category_id = matches[-1].category_id
            if category_id:
                selected_ids.append(category_id)
        if selected_ids:
            choice_values["（必須）カテゴリー"] = " | ".join(selected_ids)
    if "地場産品類型" in workbook.sheetnames:
        local_sheet = workbook["地場産品類型"]
        local_code = normalize(local_sheet["A2"].value)
        local_reason = normalize(local_sheet["B2"].value)
        if local_code or local_reason:
            choice_values["地場産品類型"] = f"{local_code}|{local_reason}".strip("|")

    def detail_rows(sheet_name: str) -> list[dict[str, str]]:
        if sheet_name not in workbook.sheetnames:
            return []
        sheet = workbook[sheet_name]
        headers = [normalize(cell.value) for cell in sheet[1]]
        result = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            values = {header: normalize(value) for header, value in zip(headers, row) if header}
            if any(value for key, value in values.items() if key != "お届け回"):
                result.append(values)
        return result

    return RegistrationExcelImport(
        product_shape=product_shape if product_shape in PRODUCT_SHAPES else "単品",
        choice_values=choice_values,
        extra_values=extra_values,
        subscription_rows=detail_rows("定期便明細"),
        sku_rows=detail_rows("SKU明細"),
        warnings=tuple(dict.fromkeys(warnings)),
    )


def _read_legacy_workbook(workbook, fields: Iterable[RequestFormField]) -> RegistrationExcelImport:
    """配置が多少異なる従来申請書を、セル位置ではなく日本語見出しで読む。"""

    sheets = [sheet for sheet in workbook.worksheets if sheet.max_row and sheet.max_column]

    def text(value) -> str:
        return normalize(value).replace("\n", " ").strip()

    def find_cells(*labels: str):
        normalized = tuple(text(label) for label in labels)
        for sheet in sheets:
            for row in sheet.iter_rows():
                for cell in row:
                    cell_text = text(cell.value)
                    if cell_text and any(label == cell_text or label in cell_text for label in normalized):
                        yield sheet, cell

    def values_right(*labels: str, limit: int = 24) -> list[str]:
        for sheet, cell in find_cells(*labels):
            values = []
            for column in range(cell.column + 1, min(sheet.max_column, cell.column + limit) + 1):
                value = text(sheet.cell(cell.row, column).value)
                if value and value not in values:
                    values.append(value)
            if values:
                return values
        return []

    def first(*labels: str) -> str:
        values = values_right(*labels)
        return values[0] if values else ""

    aliases = {
        "site business": ("事業者名", "事業者名称"),
        "管理コード": ("品番（リンベル記入）", "品番", "商品コード"),
        "（必須）お礼の品名": ("商品名（サイト掲載用）", "商品名"),
        "（条件付き必須）必要寄付金額": ("寄附額", "寄付額"),
        "容量": ("内容量・規格", "内容量", "規格"),
        "キャッチコピー": ("キャッチコピー",),
        "説明": ("商品説明", "説明文"),
        "消費期限": ("賞味期限・消費期限", "消費期限", "賞味期限"),
        "（必須）配送業者": ("配送業者",),
        "申込期日": ("申込期日", "受付期間"),
        "発送期日": ("発送期日", "納期"),
        "配送不可地域": ("配送不可地域", "配送除外地域"),
    }
    valid_columns = {field.source_column for field in fields}
    choice_values = {
        column: first(*labels)
        for column, labels in aliases.items()
        if column in valid_columns and first(*labels)
    }
    extras = {}
    for label, aliases_ in {
        "商品代（税込）": ("商品代（税込み）", "商品代（税込）", "商品代"),
        "在庫数": ("在庫数",),
        "原材料名": ("原材料名", "原材料"),
        "参考URL・SNS": ("参考URL", "SNS"),
    }.items():
        value = first(*aliases_)
        if value:
            extras[label] = value

    temperature = first("配送温度帯", "温度帯")
    if temperature:
        for label, column in {
            "常温": "常温配送フラグ", "冷蔵": "冷蔵配送フラグ", "冷凍": "冷凍配送フラグ"
        }.items():
            if column in valid_columns:
                choice_values[column] = "1" if label in temperature else "0"

    categories = load_choice_categories()
    category_values = values_right("カテゴリー", "カテゴリ", limit=30)
    selected_ids = []
    for value in category_values:
        matches = [item for item in categories if value in {item.major, item.middle, item.minor}]
        if matches:
            category_id = matches[-1].category_id
            if category_id not in selected_ids:
                selected_ids.append(category_id)
    if selected_ids and "（必須）カテゴリー" in valid_columns:
        choice_values["（必須）カテゴリー"] = " | ".join(selected_ids[:3])

    standards = dict(load_local_product_standards())
    local_values = values_right("地場産品類型", "地場産品基準", limit=30)
    if local_values and "地場産品類型" in valid_columns:
        code = next((value for value in local_values if value in standards), local_values[0])
        reason = next((value for value in local_values if value != code), "")
        choice_values["地場産品類型"] = f"{code}|{reason}".strip("|")

    if not choice_values and not extras:
        raise ValueError("入力項目を認識できませんでした。公式テンプレートまたは見出し付き申請書を使用してください。")
    return RegistrationExcelImport(
        product_shape="単品",
        choice_values=choice_values,
        extra_values=extras,
        subscription_rows=[],
        sku_rows=[],
        warnings=("従来様式を日本語見出しから読み取りました。取込後の値をご確認ください。",),
    )
