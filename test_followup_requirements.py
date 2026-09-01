from dataclasses import replace
from datetime import datetime
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from backlog_custom_fields import (
    BacklogCustomField,
    BacklogCustomFieldOption,
    build_custom_field_parameters,
    load_backlog_custom_fields,
)
from backlog_config import build_backlog_configs
from product_requests import (
    ProductCorrectionLine,
    ProductCorrectionRequest,
    ProductReference,
    SavedProductCorrectionDetail,
    SavedProductCorrectionRequest,
    build_saved_product_correction_request_summaries,
    build_backlog_issue_content,
)
from registration_excel import build_registration_template, read_registration_template
from form_definitions import RequestFormField
from request_form import (
    COMPOUND_STOCK_OPTION,
    _format_input_error_message,
    _load_saved_request_into_draft,
    _product_draft_key,
)
import streamlit as st


def _request():
    return ProductCorrectionRequest(
        request_id="REQ-1", requested_at=datetime.now(), requester="担当者",
        municipality_id="m1", municipality_name="自治体",
        note="優先して対応してください。\n\n【関連ファイル】\nBOX URL：https://example.test",
    )


def _product():
    return ProductReference(
        municipality_id="m1", municipality_name="自治体", product_id="p1",
        original_product_id="", product_name="商品", business_id="b1", business_name="事業者",
        source_row=(("管理コード", "CODE-1"), ("（必須）表示有無", "1"), ("登録年度", "2025")),
    )


def test_backlog_uses_per_product_table_and_hides_automatic_fields():
    product = _product()
    lines = [
        ProductCorrectionLine(
            product, "説明", "旧", "新", display_name="商品説明",
            image_instruction="2枚目を差し替え",
        ),
        ProductCorrectionLine(product, "登録年度", "2025", "2026", display_name="登録年度"),
        ProductCorrectionLine(product, "（必須）表示有無", "1", "0", instruction="【自動設定】"),
    ]
    _, description = build_backlog_issue_content(_request(), lines)
    assert "| 変更項目 | 現在値 | 更新後 |" in description
    assert "| 商品説明 | 旧 | 新 |" in description
    assert "登録年度" not in description
    assert "表示有無" not in description
    assert "BOX URL" in description
    assert "画像修正指示：2枚目を差し替え" in description
    assert "画像修正指示：\n" not in description
    assert "| 品番 | CODE-1 | CODE-1 |" in description
    assert "| 商品名 | 商品 | 商品 |" in description


def test_new_product_backlog_uses_compact_table_and_aggregates_choices():
    product = ProductReference(
        municipality_id="m1", municipality_name="自治体", product_id="",
        original_product_id="", product_name="新商品 NEW-1",
        business_id="", business_name="既存事業者",
        source_row=(("管理コード", ""), ("（必須）お礼の品名", "")),
    )
    request = replace(_request(), request_unit="新規商品登録", work_category="新規商品登録")
    lines = [
        ProductCorrectionLine(product, "管理コード", "", "NEW-1", display_name="品番"),
        ProductCorrectionLine(product, "（必須）お礼の品名", "", "新商品 NEW-1", display_name="商品名"),
        ProductCorrectionLine(product, "サイト表示事業者名", "", "既存事業者", display_name="事業者名"),
        ProductCorrectionLine(product, "説明", "", "商品の説明", display_name="商品説明文"),
        ProductCorrectionLine(product, "（必須）冷蔵配送", "", "1", display_name="冷蔵配送"),
        ProductCorrectionLine(product, "アレルギー：卵", "", "1", display_name="アレルギー：卵"),
        ProductCorrectionLine(product, "【Backlogのみ】税率", "", "8%", display_name="税率"),
        ProductCorrectionLine(product, "【Backlogのみ】在庫数", "", "20", display_name="在庫数"),
    ]

    _, description = build_backlog_issue_content(request, lines)

    assert "| 変更項目 | 現在値 | 更新後 |" in description
    assert "| 品番 | （未設定） | NEW-1 |" in description
    assert "| 商品名 | （未設定） | 新商品 NEW-1 |" in description
    assert "| 事業者名 | （未設定） | 既存事業者 |" in description
    assert "| 温度帯 | （未設定） | 冷蔵 |" in description
    assert "| アレルギー品目 | （未設定） | 卵 |" in description
    assert "| 在庫数 | （未設定） | 20 |" in description
    assert "税率" not in description


def test_registration_template_only_contains_required_product_management_extras():
    fields = [
        RequestFormField(
            field_id="F1", visibility="対象項目選択肢", requirement="必須",
            label="商品名", source_column="（必須）お礼の品名", input_kind="テキスト",
        ),
        RequestFormField(
            field_id="F2", visibility="対象項目選択肢", requirement="必須",
            label="事業者名", source_column="サイト表示事業者名", input_kind="テキスト",
        ),
    ]
    content = build_registration_template(
        fields,
        "単品",
        choice_values={
            "（必須）お礼の品名": "テスト商品",
            "サイト表示事業者名": "既存事業者",
        },
        extra_values={"商品代（税込）": "1000", "在庫数": "20", "税率": "8%"},
    )
    imported = read_registration_template(content, fields)
    workbook = load_workbook(BytesIO(content), data_only=False)

    assert imported.extra_values == {"商品代（税込）": "1000", "在庫数": "20"}
    assert workbook["新規返礼品申込フォーム"]["H6"].value == "既存事業者"
    assert "商品管理情報" in workbook.sheetnames
    assert "共通追加情報" not in workbook.sheetnames


def test_detail_validation_errors_are_grouped_with_line_breaks():
    message = _format_input_error_message([
        "商品A | 事業者A: 商品名、発送期日",
        "商品A | 事業者A: 商品代（税込）",
        "商品B | 事業者B: 在庫数",
    ])

    assert "**商品A | 事業者A**\n- 商品名\n- 発送期日\n- 商品代（税込）" in message
    assert "**商品B | 事業者B**\n- 在庫数" in message


def test_saved_history_restores_existing_product_values_into_input_draft():
    st.session_state.clear()
    name_field = RequestFormField(
        field_id="F1", visibility="対象項目選択肢", requirement="必須",
        label="商品名", source_column="（必須）お礼の品名", input_kind="テキスト",
    )
    product = ProductReference(
        municipality_id="m1", municipality_name="自治体", product_id="p1",
        original_product_id="", product_name="旧商品 OLD-1",
        business_id="b1", business_name="事業者",
        source_row=(("管理コード", "OLD-1"), ("（必須）お礼の品名", "旧商品 OLD-1")),
    )
    saved = SavedProductCorrectionRequest(
        request_id="REQ-OLD", backlog_issue_key="TEST-10",
        municipality_id="m1", municipality_name="自治体", requester="担当者",
        request_unit="商品単位", work_category="一般業務", note="備考",
        backlog_issue_type="既存ページ修正",
        details=(
            SavedProductCorrectionDetail(
                product_id="p1", original_product_id="",
                field_name="（必須）お礼の品名", before_value="旧商品 OLD-1",
                after_value="新商品 OLD-1",
            ),
            SavedProductCorrectionDetail(
                product_id="p1", original_product_id="",
                field_name="【Backlogのみ】在庫数", before_value="", after_value="30",
            ),
        ),
    )

    missing = _load_saved_request_into_draft(
        saved_request=saved, products=[product], form_fields=[name_field]
    )

    assert missing == []
    assert st.session_state.correction_lines == []
    assert st.session_state.loaded_history_initial_values[_product_draft_key(product)][
        "（必須）お礼の品名"
    ] == "新商品 OLD-1"
    assert st.session_state.loaded_history_backlog_values[_product_draft_key(product)][
        "在庫数"
    ] == "30"
    assert name_field in st.session_state.request_selected_form_fields
    assert COMPOUND_STOCK_OPTION in st.session_state.request_selected_form_fields
    assert st.session_state.backlog_history_action == "既存のBacklog課題を更新する"


def test_saved_new_product_without_master_id_restores_as_new_product_draft():
    st.session_state.clear()
    fields = [
        RequestFormField(
            field_id="F1", visibility="対象項目選択肢", requirement="必須",
            label="商品名", source_column="（必須）お礼の品名", input_kind="テキスト",
        ),
        RequestFormField(
            field_id="F2", visibility="対象項目選択肢", requirement="必須",
            label="事業者名", source_column="サイト表示事業者名", input_kind="テキスト",
        ),
    ]
    saved = SavedProductCorrectionRequest(
        request_id="REQ-NEW", backlog_issue_key="",
        municipality_id="m1", municipality_name="自治体", requester="担当者",
        request_unit="新規商品登録", work_category="新規商品登録", note="備考",
        backlog_issue_type="新規商品登録",
        details=(
            SavedProductCorrectionDetail(
                product_id="", original_product_id="",
                field_name="（必須）お礼の品名", before_value="", after_value="新商品 NEW-1",
            ),
            SavedProductCorrectionDetail(
                product_id="", original_product_id="",
                field_name="サイト表示事業者名", before_value="", after_value="既存事業者",
            ),
            SavedProductCorrectionDetail(
                product_id="", original_product_id="",
                field_name="【Backlogのみ】商品代（税込）", before_value="", after_value="1200",
            ),
        ),
    )

    missing = _load_saved_request_into_draft(
        saved_request=saved, products=[], form_fields=fields
    )

    assert missing == []
    assert st.session_state.request_mode == "新規商品登録"
    assert st.session_state.loaded_history_new_product_values["（必須）お礼の品名"] == "新商品 NEW-1"
    assert st.session_state.loaded_history_business_name == "既存事業者"
    assert st.session_state.new_product_backlog_values["商品代（税込）"] == "1200"
    assert st.session_state.backlog_history_action == "新しいBacklog課題を作成する"


def test_history_edit_can_choose_existing_issue_update_or_new_issue_creation():
    source = Path(__file__).with_name("request_form.py").read_text(encoding="utf-8")

    assert '"既存のBacklog課題を更新する"' in source
    assert '"新しいBacklog課題を作成する"' in source
    assert 'key="backlog_history_action"' in source
    assert "backlog_issue_key_to_update = (" in source
    update_branch = source.index("if backlog_issue_key_to_update:")
    create_branch = source.index("else:\n                    issue = create_issue(", update_branch)
    assert update_branch < create_branch


def test_backlog_config_accepts_product_code_routing_columns():
    config = build_backlog_configs([{
        "連携有効": "1", "自治体ID": "m1", "自治体名": "自治体",
        "BacklogスペースID": "space", "BacklogプロジェクトID": "1",
        "Backlog APIキー": "secret", "品番担当者ユーザーID": "10",
        "品番通知先ユーザーID": "10|20",
    }])[0]
    assert config.product_code_assignee_id == "10"
    assert config.product_code_notified_user_ids == ("10", "20")


def test_page_correction_template_and_table_use_business_priority_order():
    product = ProductReference(
        municipality_id="m1", municipality_name="自治体", product_id="p1",
        original_product_id="", product_name="旧商品 OLD-1", business_id="b1", business_name="事業者",
        source_row=(("管理コード", "OLD-1"), ("（条件付き必須）必要寄付金額", "10000")),
    )
    request = replace(_request(), request_unit="商品単位")
    lines = [
        ProductCorrectionLine(product, "説明", "旧説明", "新説明", display_name="商品説明"),
        ProductCorrectionLine(product, "【Backlogのみ】在庫数", "", "20", display_name="在庫数"),
        ProductCorrectionLine(product, "【Backlogのみ】商品代（税込）", "1200", "1400", display_name="商品代"),
        ProductCorrectionLine(product, "（条件付き必須）必要寄付金額", "10000", "12000", display_name="寄付額"),
        ProductCorrectionLine(product, "（必須）お礼の品名", "旧商品 OLD-1", "新商品 NEW-1", display_name="商品名"),
        ProductCorrectionLine(product, "管理コード", "OLD-1", "NEW-1", display_name="品番"),
    ]
    _, description = build_backlog_issue_content(request, lines)
    assert "【ページ修正テンプレート】" not in description
    assert "【商品の変更点】" not in description
    assert "■ 商品" not in description
    assert description.index("対応内容・備考：") < description.index("| 品番 | OLD-1 | NEW-1 |")
    assert "対応内容・備考：優先して対応してください。" in description
    assert "対応内容・備考：\n" not in description
    assert "品番：OLD-1 → NEW-1" not in description
    assert "商品名：旧商品 OLD-1 → 新商品 NEW-1" not in description
    assert "事業者名：事業者" not in description
    assert "| 品番 | OLD-1 | NEW-1 |" in description
    assert "| 商品名 | 旧商品 OLD-1 | 新商品 NEW-1 |" in description
    assert "| 寄附額 | 10000 | 12000 |" in description
    assert "| 在庫数 | （未設定） | 20 |" in description
    assert "| 商品代 | — | 1400（変更後商品代） |" in description
    assert description.rstrip().endswith("BOX URL：https://example.test")
    assert description.index("| 品番 |") < description.index("| 商品名 |")
    assert description.index("| 商品名 |") < description.index("| 寄附額 |")
    assert description.index("| 寄附額 |") < description.index("| 商品代 |")
    assert description.index("| 商品代 |") < description.index("| 在庫数 |")


def test_product_code_request_shows_current_code_and_blank_instruction():
    product = ProductReference(
        municipality_id="m1", municipality_name="自治体", product_id="p1",
        original_product_id="", product_name="商品 OLD-1",
        business_id="b1", business_name="事業者",
        source_row=(("管理コード", "OLD-1"),),
    )
    request = replace(_request(), request_unit="商品単位")
    lines = [ProductCorrectionLine(
        product, "【Backlogのみ】品番取得依頼", "OLD-1",
        "（空欄：品番を取得してください）", display_name="品番",
    )]
    _, description = build_backlog_issue_content(request, lines)
    assert "品番：OLD-1 → （空欄：品番を取得してください）" not in description
    assert "商品名：商品 OLD-1 → 商品" not in description
    assert "| 品番 | OLD-1 | （空欄：品番を取得してください） |" in description
    assert "| 商品名 | 商品 OLD-1 | 商品 |" in description
    assert "【商品の変更点】" not in description


def test_duplicate_product_name_lines_are_merged_and_use_new_code():
    product = ProductReference(
        municipality_id="m1", municipality_name="自治体", product_id="p1",
        original_product_id="", product_name="旧商品 OLD-1",
        business_id="b1", business_name="事業者",
        source_row=(("管理コード", "OLD-1"),),
    )
    request = replace(_request(), request_unit="商品単位")
    lines = [
        ProductCorrectionLine(
            product, "（必須）お礼の品名", "旧商品 OLD-1", "新しい商品 OLD-1",
            display_name="商品名",
        ),
        ProductCorrectionLine(
            product, "（必須）お礼の品名", "旧商品 OLD-1", "旧商品 NEW-1",
            display_name="商品名",
        ),
        ProductCorrectionLine(product, "管理コード", "OLD-1", "NEW-1", display_name="品番"),
    ]
    _, description = build_backlog_issue_content(request, lines)
    assert description.count("| 商品名 |") == 1
    assert "商品名：旧商品 OLD-1 → 新しい商品 NEW-1" not in description
    assert "| 商品名 | 旧商品 OLD-1 | 新しい商品 NEW-1 |" in description


def test_donation_correction_types_control_product_cost_input():
    source = __import__("pathlib").Path(__file__).with_name("request_form.py").read_text(
        encoding="utf-8"
    )
    assert '"寄附額変更（商品代変更アリ）"' in source
    assert '"寄附額変更（商品代変更ナシ）"' in source
    cost_section = source.index('st.subheader("商品代の変更")')
    cost_condition = source.rfind("if donation_with_cost_requested:", 0, cost_section)
    assert cost_condition >= 0
    no_cost_branch = source.index("elif donation_without_cost_requested:", cost_section)
    assert '"商品代変更": "商品代を変更しない"' in source[
        no_cost_branch:no_cost_branch + 400
    ]


def test_compound_donation_options_control_product_cost_input():
    source = __import__("pathlib").Path(__file__).with_name("request_form.py").read_text(
        encoding="utf-8"
    )
    assert 'COMPOUND_DONATION_WITH_COST_OPTION = "寄附額（商品代変更アリ）"' in source
    assert 'COMPOUND_DONATION_WITHOUT_COST_OPTION = "寄附額（商品代変更ナシ）"' in source
    assert "and COMPOUND_DONATION_WITH_COST_OPTION in selected_change_options" in source
    assert "and COMPOUND_DONATION_WITHOUT_COST_OPTION in selected_change_options" in source
    assert "if donation_with_cost_requested:" in source
    assert "and not _is_donation_field(field)" in source


def test_donation_field_detection_handles_master_label_variants():
    source = __import__("pathlib").Path(__file__).with_name("request_form.py").read_text(
        encoding="utf-8"
    )
    helper = source[source.index("def _is_donation_field"):source.index("def _sort_change_fields")]
    assert 'field.source_column == POINTS_COLUMN' in helper
    assert '("必要寄付金額", "寄附額", "寄付額")' in helper


def test_initial_request_flow_is_limited_to_correction_or_new_registration():
    source = __import__("pathlib").Path(__file__).with_name("request_form.py").read_text(
        encoding="utf-8"
    )
    assert 'REQUEST_MODES = ("修正", "新規商品登録")' in source
    assert '"依頼内容",\n                REQUEST_MODES' in source
    assert 'request_unit = "商品単位"' in source
    assert 'work_category = "新規商品登録" if request_mode == "新規商品登録" else "一般業務"' in source
    assert '"定期便・SKU展開を利用する（試験運用）"' in source
    assert 'product_shape = "単品"' in source


def test_new_product_shapes_initialize_correction_only_state():
    source = __import__("pathlib").Path(__file__).with_name("request_form.py").read_text(
        encoding="utf-8"
    )
    initialization = source.index('correction_type = ""')
    new_product_branch = source.index("if is_new_product:", initialization)
    donation_condition = source.index("donation_with_cost_requested = (", new_product_branch)
    assert initialization < new_product_branch < donation_condition
    assert "selected_change_options: list[RequestFormField | str] = []" in source[
        initialization:new_product_branch
    ]


def test_custom_field_loader_accepts_string_credential_path():
    missing_credentials = str(Path(__file__).with_name("missing-service-account.json"))

    with pytest.raises(FileNotFoundError):
        load_backlog_custom_fields("spreadsheet-id", missing_credentials)


def test_multi_value_custom_field_uses_backlog_issue_api_parameter_name():
    field = BacklogCustomField(
        municipality_id="kaga",
        municipality_name="石川県加賀市",
        project_id="1",
        issue_type_name="既存ページ修正",
        issue_type_id="2",
        name="既存ページ修正内容",
        field_id="324256",
        type_id="7",
        required=True,
        options=(
            BacklogCustomFieldOption(name="商品情報", option_id="101"),
            BacklogCustomFieldOption(name="画像", option_id="102"),
        ),
    )

    parameters = build_custom_field_parameters(
        [field], {"既存ページ修正内容": ["商品情報", "画像"]}
    )

    assert parameters == {"customField_324256": ["101", "102"]}
    assert "customField_324256[]" not in parameters


def test_failed_backlog_request_retry_does_not_save_duplicate_master_rows():
    source = Path(__file__).with_name("request_form.py").read_text(encoding="utf-8")

    assert "retrying_failed_backlog = bool(" in source
    retry_branch = source.index("if retrying_failed_backlog:")
    normal_save = source.index("result = save_product_correction_request(", retry_branch)
    assert "request = replace(request, request_id=editing_source_request_id)" in source[
        retry_branch:normal_save
    ]
    assert "else:" in source[retry_branch:normal_save]
    assert "st.session_state.editing_source_request_id = result.request_id" in source


def test_backlog_due_date_defaults_follow_request_type_and_business_days():
    from datetime import date

    from request_form import _default_backlog_due_date

    assert _default_backlog_due_date(
        request_date=date(2026, 8, 24),
        is_new_product=False,
        correction_type="在庫数変更",
    ) == date(2026, 8, 24)
    assert _default_backlog_due_date(
        request_date=date(2026, 8, 28),
        is_new_product=False,
        correction_type="複合的な修正",
    ) == date(2026, 8, 31)
    assert _default_backlog_due_date(
        request_date=date(2026, 8, 28),
        is_new_product=True,
        correction_type="",
    ) == date(2026, 9, 2)
    # 11/23（勤労感謝の日）に当たるため翌営業日へ送る。
    assert _default_backlog_due_date(
        request_date=date(2026, 11, 20),
        is_new_product=False,
        correction_type="複合的な修正",
    ) == date(2026, 11, 24)


def test_logged_in_user_is_limited_to_authorized_municipalities():
    from backlog_users import BacklogProjectUser
    from request_form import _allowed_municipality_ids

    users = [
        BacklogProjectUser(
            municipality_id="kaga",
            municipality_name="加賀市",
            project_id="1",
            user_id="10",
            name="担当A",
            mail_address="user@example.com",
            login_address="user@example.com",
        ),
        BacklogProjectUser(
            municipality_id="ebetsu",
            municipality_name="江別市",
            project_id="2",
            user_id="11",
            name="担当A",
            mail_address="user@example.com",
            login_address="USER@example.com",
        ),
        BacklogProjectUser(
            municipality_id="kuma",
            municipality_name="球磨村",
            project_id="3",
            user_id="12",
            name="担当B",
            mail_address="other@example.com",
            login_address="other@example.com",
        ),
    ]

    assert _allowed_municipality_ids(users, "User@Example.com") == {"kaga", "ebetsu"}


def test_urgent_request_prefixes_backlog_issue_title():
    source = Path(__file__).with_name("request_form.py").read_text(encoding="utf-8")

    assert 'st.checkbox(\n                "至急"' in source
    assert 'issue_summary = f"【至急】{issue_summary}"' in source


def test_municipality_sort_order_runs_from_hokkaido_to_okinawa():
    from request_form import _municipality_sort_key

    names = ["沖縄県那覇市", "福島県相馬市", "北海道江別市", "青森県平内町"]

    assert sorted(names, key=_municipality_sort_key) == [
        "北海道江別市", "青森県平内町", "福島県相馬市", "沖縄県那覇市"
    ]


def test_public_product_filter_uses_choice_display_flag_one():
    from product_requests import ProductReference
    from request_form import _is_public_product

    def product(display_value: str) -> ProductReference:
        return ProductReference(
            municipality_id="sample",
            municipality_name="北海道サンプル町",
            product_id="1",
            original_product_id="1",
            product_name="商品",
            business_id="1",
            business_name="事業者",
            source_row=(("（必須）表示有無", display_value),),
        )

    assert _is_public_product(product("1")) is True
    assert _is_public_product(product("0")) is False
    assert _is_public_product(product("")) is False


def test_saved_request_history_is_permission_filtered_and_newest_first():
    rows = [
        {
            "依頼ID": "PR-OLD",
            "依頼日時": "2026-08-20 09:00:00",
            "依頼者": "担当A",
            "自治体ID": "allowed",
            "自治体名": "北海道許可町",
            "依頼種別": "商品修正",
            "Backlog親課題キー": "TEST-1",
            "Backlog親課題URL": "https://example.test/TEST-1",
            "状態": "受付",
        },
        {
            "依頼ID": "PR-NEW",
            "依頼日時": "2026-08-24 10:00:00",
            "依頼者": "担当B",
            "自治体ID": "allowed",
            "自治体名": "北海道許可町",
            "依頼種別": "商品修正",
            "Backlog親課題キー": "",
            "状態": "Backlog起票待ち",
        },
        {
            "依頼ID": "PR-HIDDEN",
            "依頼日時": "2026-08-25 10:00:00",
            "依頼者": "担当C",
            "自治体ID": "forbidden",
            "自治体名": "権限外市",
        },
    ]

    summaries = build_saved_product_correction_request_summaries(
        rows, allowed_municipality_ids={"allowed"}
    )

    assert [summary.request_id for summary in summaries] == ["PR-NEW", "PR-OLD"]
    assert summaries[0].lookup_value == "PR-NEW"
    assert summaries[1].lookup_value == "TEST-1"
    assert all(summary.municipality_id == "allowed" for summary in summaries)


def test_history_screen_explains_identifiers_and_supports_other_requesters():
    source = Path(__file__).with_name("request_form.py").read_text(encoding="utf-8")

    assert '("自分の依頼", "閲覧可能な自治体の全依頼")' in source
    assert "再編集したい依頼の行をクリックして選択してください。" in source
    assert "Backlog課題キー**" in source
    assert "商品マスタの「商品修正依頼」シート" in source
    assert "まだ「依頼を保存」を押していない入力途中の内容は表示されません。" in source
    assert "allowed_municipality_ids=allowed_municipality_ids" in source


def test_saved_request_history_is_collapsed_by_default():
    source = Path(__file__).with_name("request_form.py").read_text(encoding="utf-8")

    history_start = source.index('"登録済み依頼を再編集",')
    history_section = source[history_start:history_start + 220]
    assert "with st.expander(" in source[history_start - 40:history_start]
    assert "expanded=False" in history_section
    assert 'icon=":material/history:"' in history_section



def test_dedicated_correction_types_always_resolve_their_input_fields():
    source = __import__("pathlib").Path(__file__).with_name("request_form.py").read_text(
        encoding="utf-8"
    )
    helper = source[
        source.index("def _fields_for_correction_type"):
        source.index("def _field_visibility")
    ]
    assert "CORRECTION_TYPE_COLUMNS.get(correction_type, set())" in helper
    assert "field.source_column in target_columns" in helper
    assert "field.source_column not in HIDDEN_FORM_COLUMNS" in helper
    assert "_field_visibility" not in helper
    assert "_fields_for_correction_type(\n                    form_fields, correction_type" in source


def test_existing_product_change_fields_start_from_current_values():
    source = __import__("pathlib").Path(__file__).with_name("request_form.py").read_text(
        encoding="utf-8"
    )
    editor = source[
        source.index("def _render_product_change_editor"):
        source.index("def _build_product_change_lines")
    ]
    assert "if field.source_column in initial_values:" in editor
    assert "elif product.product_id:" in editor
    assert 'initial_value = source_values.get(field.source_column, "")' in editor
    assert "if field.source_column == SHIPPING_DEADLINE_COLUMN:" in editor



def test_shipping_deadline_uses_choice_codes_and_free_text_only_for_custom_mode():
    source = __import__("pathlib").Path(__file__).with_name("request_form.py").read_text(
        encoding="utf-8"
    )
    assert 'deadline_value not in {"0", "任意入力"}' in source
    assert 'if selected_label == "任意入力":' in source
    assert '選択したチョイス固定値で更新します' in source


def test_request_master_cache_does_not_expire_during_form_entry():
    source = __import__("pathlib").Path(__file__).with_name("request_form.py").read_text(
        encoding="utf-8"
    )
    cache_section = source[
        source.index("def _load_products") - 80:
        source.index("def _product_label")
    ]
    assert "ttl=" not in cache_section
    assert cache_section.count("show_spinner=False") >= 9
    assert 'st.session_state[oauth_session_key] = True' in source
    assert '商品・自治体・設定マスタを再読み込み' in source
